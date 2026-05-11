import asyncio
import argparse
import datetime
import json
import os
import io
import contextlib
import copy
import time
import threading
from openpyxl import load_workbook
from aiofiles import open as aio_open
from aiofiles.os import remove as aio_remove
from aiofiles.os import path as aio_path
from aiofiles.os import makedirs as aio_makedirs
from llm.agent.symbol import Symbol
from llm.agent.errorFixer import ErrorFixer
from llm.base import DATATYPE, MODEL, K_NUM, Temperature
from utils.eval import check_exp, get_exception
from utils.exp import fix_exps
from multiprocessing import Pool, Manager, Lock
from utils.sat_problem_solver import LSAT_Z3_Program
from utils.others import clean_output, periodic_flush, get_rate, play_sound

# Parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument("K_NUM", type=int, help="Number of examples")
parser.add_argument("Temperature", type=float, help="Temperature for the model")
# adding one more parser argument for N_SAMPLES
parser.add_argument("--n_samples", type=int, default=None)

args = parser.parse_args()
# K_NUM = args.K_NUM
# Temperature = args.Temperature
Error_NUM = args.K_NUM
error_temperature = args.Temperature
# Error_NUM = 0
# error_temperature = 0

# Redo experiment
reDo = True
FIX_FLAG = True

set_name = "dev"
if set_name != "Test data":
    input_name = f"./data/{DATATYPE}-{set_name}.jsonl"
else:
    input_name = f"./data/{set_name}.jsonl"

name = f"{set_name}_{K_NUM}_{Temperature}-{Error_NUM}_{error_temperature}"

if FIX_FLAG:
    name = name + "F"

if "/" in MODEL:
    out_path = f"./data/{DATATYPE}/{MODEL.split('/')[-1]}"
else:
    out_path = f"./data/{DATATYPE}/{MODEL}"
if not os.path.exists(out_path):
    os.makedirs(out_path)

# Path for recording print output
txt_dir = f"{out_path}/{name}_outputrecord/log"
if not os.path.exists(txt_dir):
    os.makedirs(txt_dir)

output_excel = f"{out_path}/{name}.xlsx"
output_jsonl = f"{out_path}/{name}.jsonl"
template_file = "./data/symbol.xlsx"
from runner import NUM_PROCESSES
max_workers = NUM_PROCESSES


symbol_ai = Symbol(Temperature,K_NUM,5)  # Instantiate Symbol object
fixer_ai = ErrorFixer(Error_NUM, error_temperature)
# fixer_ai = ErrorFixer()


def process_normal(line,lock, processed_ids):
    data = json.loads(line)
    if data["id"] in processed_ids:
        # print(f"skip processed row: {data['id']}")
        return

    print("Start execution:", data["id"], " time:", datetime.datetime.now())
    response_text = "Exception, did not enter code generation"
    context_text = data["context"]
    question_text = data["question"]
    options_text = "\n".join(data["options"])
    input_nl = f"Nature language:\nContext:\n{context_text}\nQuestion:\n{question_text}\nOptions:\n{options_text}\n"
    actual_logic_program = ["Exception, no program generated"]       # Record current program
    temp_logic_program = []                           # Record temporary logic program for easy saving and checking
    last_error_messages = "no syntax problem"               # Record last error message
    other_errors = ""

    try:
        actual_answer = "Error"     # Initialize answer as error
        reloadtimes = 0
        MAX_RELOAD = 3              # Maximum number of times problem is allowed to be generated

        while actual_answer in ["Error", "N", "X"] and reloadtimes < MAX_RELOAD:
            reloadtimes += 1

            print(f"\n=============== ***{data['id']}: Generate program -> attempt {reloadtimes}*** ===============")
            # region
            prompt = f"Input:\nContext:\n{context_text}\nQuestion:\n{question_text}\nOptions:\n{options_text}"
            result_json, response_text = asyncio.run(symbol_ai.chat(prompt))
            # Check program, only put in actual_logic_program if there are no format issues
            temp_logic_program = result_json.get("raw_logic_programs", result_json.get("raw_logic_program", ["Exception, no program generated"]))
            if isinstance(temp_logic_program, str):
                temp_logic_program = [temp_logic_program]
            if all(substring in temp_logic_program[0] for substring in ["# Declaration", "# Constraints", "# Options"]):
                actual_logic_program = copy.deepcopy(temp_logic_program)

            response_text_print = response_text.replace('```', '`\\``').replace('\n', '\\n')
            print(f"# The {reloadtimes}th generation answer\n> {response_text_print}\n")
            print(f"## This generation program:\n```\n{input_nl}\n\n{actual_logic_program[0]}\n```\n")
            # Try to get the initial answer of this generation for easy comparison after subsequent repairs
            generate_output = None
            original_parsing_error = None
            try:
                z3_program = LSAT_Z3_Program(actual_logic_program[0], 'AR-LSAT', data['id'])
                generate_output, original_parsing_error = z3_program.execute_program()
                original_answer = z3_program.answer_mapping(generate_output)
            except Exception as e:
                original_answer = "Error"
            # endregion
            print(f"- This generation program answer:{original_answer}\n")
            print(f"=== ***The {reloadtimes}th generation complete*** ===\n")
            # generation complete

            # # Fine-tune the formula. Currently mainly fixes incomplete variable issues, e.g., `[p1, p2:person]` changed to `[p1:person, p2:person]`
            if FIX_FLAG:
                actual_logic_program[0] = fix_exps(actual_logic_program[0])

            trytimes = 0
            MAX_TRY = 3                                 # Maximum 3 times allowed to fix the problem
            error_messages = "no syntax problem"              # This initialization will be overwritten later
            other_errors = ""                           # Used to record error types not included
            error_types = []                            # Record error type
            fixed_output = ""                           # Record the fixed answer
            program_output = generate_output            # Initialize to the initial program output
            parsing_error = original_parsing_error
            actual_answer = original_answer
            while FIX_FLAG and trytimes < MAX_TRY and error_messages:
                trytimes += 1

                print(f"------------- **{data['id']}: Syntax repair -> attempt {trytimes}** -------------\n")
                error_messages, error_types, other_errors = get_exception(actual_logic_program[0], data['id'])

                # enter check if there are problems
                if error_messages:
                    error_messages_print = "\n".join(f"> {line}" for line in error_messages.splitlines())
                    print(f"ID {data['id']} syntax problem:\n\n{error_messages_print}\n\n")
                    last_error_messages = error_messages

                    # Call model for repair
                    prompt = f"{input_nl}\nSAT logic program:\n```plaintext\n{actual_logic_program[0]}\n```\nError message:{error_messages}\n"
                    fixed_json, fixed_output = asyncio.run(fixer_ai.chat(prompt, error_types))
                    fixed_output_print = fixed_output.replace('```', '`\\``').replace('\n', '\\n')
                    print(f"# The {trytimes}th repair answer\n\n> {fixed_output_print}\n")

                    # ensure generated program is normal before replacing
                    temp_logic_program = fixed_json.get("raw_logic_programs", ["Exception, no program generated"])
                    if all(substring in temp_logic_program[0] for substring in ["# Declaration", "# Constraints", "# Options"]):
                        actual_logic_program = copy.deepcopy(temp_logic_program)
                    print(f"## New program after repair:\n```\n{actual_logic_program[0]}\n```\n")
                else:
                    print("- No syntax error detected")

                actual_logic_program[0] = fix_exps(actual_logic_program[0])     # Fine-tune the formula

                # get answer
                try:
                    z3_program = LSAT_Z3_Program(actual_logic_program[0], 'AR-LSAT', data['id'])
                    program_output, parsing_error = z3_program.execute_program()
                    actual_answer = z3_program.answer_mapping(program_output)
                    print(f"- Answer after repair:{actual_answer}\n")
                except Exception as e:
                    program_output = []
                    actual_answer = "Error"
                    print(f"- This solution failed. Failure reason:{str(e)}\n")

                # if answer has been selected, exit repair loop
                if actual_answer not in ["X", "N", "Error"]:
                    if original_answer in ["X", "N", "Error"]:
                        print(f"# Repair successful\n```\n-----*****-----\noriginal answer: {original_answer}\ncurrent answer: {actual_answer}\n-----*****-----\n```\n")
                    break
                # if no syntax error, should also exit loop
                if not error_messages:
                    break
                # notify that max repair attempts have been reached
                if trytimes == MAX_TRY:
                    print(f"Reached the maximum number of repairs for this round...\ncurrent answer: {actual_answer}")

                # Reinitialize error message to prepare for next repair
                error_messages = "no syntax problem"

        # If you want to select multiple, use the first one as the answer, use the following three lines
        # if actual_answer != "Error":
        #     z3_program = LSAT_Z3_Program(actual_logic_program[0], 'AR-LSAT', data['id'])
        #     actual_answer = z3_program.answer_mapping(program_output, "final")

        # Get the last error, that is, the error in the current SAT program
        program_error = ""
        if program_output:
            program_error = program_output[4:]
        emsg = f"# Final program runtime error:\n1. Normal error:{program_error}\n2. Parsing error:{parsing_error}\n3. Unrecorded error:{other_errors}\n"
        print(f"# Final program as follows\n```\n{input_nl}\n\n{actual_logic_program[0]}\n```\n- Normal error situation:\n{program_error}\n*******\n- Parsing error:\n{parsing_error}\n")

        result_row = {
            "id": data["id"],
            "context": context_text,
            "question": question_text,
            "options": options_text,
            "logic_program": actual_logic_program[0],
            "output": response_text,
            "error_information": last_error_messages,
            "answer": data["answer"],
            "actual_answer": actual_answer,
            "error_msg": emsg
        }

    except Exception as e:
        print("Error processing:", data["id"], e)
        result_row = {
            "id": data["id"],
            "context": context_text,
            "question": question_text,
            "options": options_text,
            "logic_program": actual_logic_program[0],
            "output": response_text,
            "error_information": last_error_messages,
            "answer": data["answer"],
            "actual_answer": "Error",
            "error_msg": str(e) + f"omitted type: {other_errors}"
        }

    # use file lock to perform write operation
    try:
        with lock:
            print("Save results:", data["id"], " time:", datetime.datetime.now())
            # write JSONL file
            with open(output_jsonl, 'a', encoding='utf-8') as jsonl_file:
                jsonl_file.write(json.dumps(result_row, ensure_ascii=False) + '\n')

            # if need to write to excel, enable by uncommenting below. Also enable corresponding content in main
            # wb = load_workbook(output_excel)
            # ws = wb.active
            # ws.append([
            #     result_row["id"],
            #     result_row["context"],
            #     result_row["question"],
            #     result_row["options"],
            #     result_row["logic_program"],
            #     result_row["output"],
            #     result_row["error_information"],
            #     fixed_output,
            #     result_row["answer"],
            #     result_row["actual_answer"],
            #     result_row["answer"] == result_row["actual_answer"],
            #     result_row["error_msg"]
            # ])
            # wb.save(output_excel)
            get_rate(output_jsonl)
    except Exception as e:
        print("write error:", data["id"], e, result_row)
    print("Execution completed:", data["id"], " time:", datetime.datetime.now())


# Global shared list of ids being processed and start time dictionary
manager = None
processing_ids = None
processing_start_times = None

def process_wrapper(args):
    """Wrapper function to pass multiple parameters to the target function"""
    # Using this line will output to terminal. Speed may be a bit faster
    # return process_normal(*args)

    line, lock, processed_ids, processing_ids, processing_start_times, id_count = args

    line_data = json.loads(line)
    line_id = line_data["id"]

    # Record the start time of the current id
    start_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    # Add the current id to the processing list and record the start time
    with lock:
        processing_ids.append(line_id)
        id_count.value += 1
        processing_start_times[line_id] = start_time
        # real-time output of currently processing id list and start times
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # Get current time
        processing_info = ", ".join([f"\n{id} ({processing_start_times[id]})" for id in processing_ids])
        print(f"Processing:[{processing_info}]\n -----{current_time}: Total {id_count.value}-----\n")

    # stream object captures output
    output_capture = io.StringIO()
    output_file_name = f"{txt_dir}/{line_id}.md"

    # create exit flag
    stop_event = threading.Event()

    # start periodic_flush
    flush_thread = threading.Thread(target=periodic_flush, args=(output_capture, output_file_name, stop_event, 20))
    flush_thread.start()

    # use redirect_stdout to redirect standard output to StringIO object
    with contextlib.redirect_stdout(output_capture):
        process_normal(line, lock, processed_ids)

    stop_event.set()  # notify flush thread to exit
    flush_thread.join()  # wait for flush thread to complete

    # write remaining output to file
    captured_output = output_capture.getvalue()
    cleaned_output = clean_output(captured_output)
    if cleaned_output:
        with open(output_file_name, 'a', encoding='utf-8') as f:
            f.write(cleaned_output)

    # close StringIO object
    output_capture.close()

    # remove the current id from the processing list
    with lock:
        processing_ids.remove(line_id)
        id_count.value -= 1
        del processing_start_times[line_id]  # delete start time record
        # real-time output of currently processing id list and start times
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # Get current time
        processing_info = ",".join([f"\n{id} ({processing_start_times[id]})" for id in processing_ids])
        print(f"Processing:[{processing_info}]\n -----{current_time}: Total {id_count.value}-----")

    # print process progress
    get_rate(output_jsonl)
    print(f"{line_id} completed")


def main():
    global manager, processing_ids, processing_start_times, id_count

    # initialize Manager and shared variables
    manager = Manager()
    processing_ids = manager.list()
    processing_start_times = manager.dict()
    id_count = manager.Value('i', 0)  # use manager.Value to create shared integer, initial value is 0

    # load reference template and save to new file
    # wb = load_workbook(template_file)
    if reDo:
        # If there is an output file, delete it first
        if os.path.exists(output_excel):
            os.remove(output_excel)
        if os.path.exists(output_jsonl):
            os.remove(output_jsonl)
        # wb.save(output_excel)

    # If redo experiment, clear all previous records, print records and logic program records
    if reDo:
        for filename in os.listdir(txt_dir):
                file_path = os.path.join(txt_dir, filename)
                if file_path.endswith(".md"):
                    os.remove(file_path)

        # clear logic program folder
        logic_program_folder = "utils/.cache_program"
        if os.path.exists(logic_program_folder):
            # clear all records
            for filename in os.listdir(logic_program_folder):
                    file_path = os.path.join(logic_program_folder, filename)
                    if file_path.endswith(".py"):
                        os.remove(file_path)

    # Open JSONL file
    with open(input_name, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        # addding samples logic
        if args.n_samples is not None:
            lines = lines[:args.n_samples]


    # Read processed ID
    processed_ids = set()
    if os.path.exists(output_jsonl):
        with open(output_jsonl, 'r', encoding='utf-8') as jsonl_file:
            for line in jsonl_file:
                result_row = json.loads(line)
                processed_ids.add(result_row["id"])

    with manager:
        lock = manager.Lock()  # File lock
        pool = Pool(processes=max_workers)

        # Create parameter list to pass to process
        pool_args = [(line, lock, processed_ids, processing_ids, processing_start_times, id_count) for line in lines]
        # In parallel process
        pool.map(process_wrapper, pool_args)

        pool.close()
        pool.join()

        # record experiment result
    result = get_rate(output_jsonl)
    with open("experiment_results.txt", "a", encoding="utf-8") as f:
        f.write(f"\nMODEL: {MODEL}, K_NUM: {K_NUM}, Temperature: {Temperature}\n")
        if FIX_FLAG:
            f.write(f"repair number of examples: {Error_NUM}, repair temperature: {error_temperature}\n")
        f.write(f"{result[0]}/{result[1]}, Accuracy: {result[2]}\n")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Exception occurred:{str(e)}")

    # play_sound(duration=0.7)
    get_rate(output_jsonl)
