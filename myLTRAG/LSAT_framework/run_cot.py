import datetime
import json
import os
import io
import threading
import contextlib
from utils.others import clean_output, periodic_flush, get_rate, account_balance
from multiprocessing import Pool, Manager, Lock
from openpyxl import load_workbook
from llm.agent.cot import CoT
from llm.base import DATATYPE, MODEL
import asyncio
import argparse

#parsing command line args. Only N_SAMPLES mod for now.
parser = argparse.ArgumentParser()
parser.add_argument("--n_samples", type=int, default=None)
args = parser.parse_args()

needExample = False

set_name = "lsat-dev"
input_name = f"./data/{set_name}.jsonl"

name = "Standard"
if "/" in MODEL:
    out_path = f"./data/{DATATYPE}/{MODEL.split('/')[-1]}"
else:
    out_path = f"./data/{DATATYPE}/{MODEL}"
if not os.path.exists(out_path):
    os.makedirs(out_path)

# Path for recording output
txt_dir = f"{out_path}/Standard record"
# region
if not os.path.exists(txt_dir):
    os.makedirs(txt_dir)
# Clear all previous records
for filename in os.listdir(txt_dir):
        file_path = os.path.join(txt_dir, filename)
        if file_path.endswith(".txt"):
            os.remove(file_path)
# endregion

output_name = f"{out_path}/{name}.xlsx"
output_jsonl = f"{out_path}/{name}.jsonl"
template_file = "./data/cot.xlsx"
from runner import NUM_PROCESSES
max_workers = NUM_PROCESSES

CoTer = CoT(needExample)  # Instantiate Cot object

def process_normal(line, lock):
    data = json.loads(line)
    print("Start execution:", data["id"], " time:", datetime.datetime.now())
    context_text = data["context"]
    question_text = data["question"]
    options_text = "\n".join(data["options"])
    prompt = f"Context:\n{context_text}\nQuestion:\n{question_text}\nOptions:\n{options_text}\n"
    # Call cot_agent's chat method to get results
    result_json, response_text = asyncio.run(CoTer.chat(prompt))
    print(f"This answer:\n{response_text}\n\n\n")
    actual_answer = result_json.get("answer", "Error")
    print(f"answer = {data['answer']}\nanctual answer = {actual_answer}\n")

    result_row = [
        data["id"],
        context_text,
        question_text,
        options_text,
        response_text,
        data.get("answer", ""),
        actual_answer,
        actual_answer == data.get("answer", "")
    ]

    result_dict = {
        "id": data["id"],
        "context": context_text,
        "question": question_text,
        "options": options_text,
        "output": "See folder ‘CoT record’ with corresponding id .txt file",
        "answer": data.get("answer", ""),
        "actual_answer": actual_answer,
        "error_msg": ""  # Error message placeholder if needed
    }

    # UseFile lockperform write operation
    with lock:
        print("Save results ", data["id"], " time：", datetime.datetime.now())
        # wb = load_workbook(output_name)
        # ws = wb.active
        # ws.append(result_row)
        # wb.save(output_name)

        # write JSONL file
        with open(output_jsonl, 'a', encoding='utf-8') as jsonl_file:
            jsonl_file.write(json.dumps(result_dict, ensure_ascii=False) + '\n')
    get_rate(output_jsonl)

def process_wrapper(args):
    """Wrapper function to pass multiple parameters to the target function"""
    # Using this line will output to terminal. Speed may be a bit faster
    # return process_normal(*args)
    
    line, lock = args

    line_data = json.loads(line)
    line_id = line_data["id"]

    # stream object captures output
    output_capture = io.StringIO()
    output_file_name = f"{txt_dir}/{line_id}.txt"

    # create exit flag
    stop_event = threading.Event()

    # startperiodic_flush
    flush_thread = threading.Thread(target=periodic_flush, args=(output_capture, output_file_name, stop_event, 10))
    flush_thread.start()

    # Use redirect_stdout convert standardoutputredirect to StringIO object
    with contextlib.redirect_stdout(output_capture):
        process_normal(line, lock)

    stop_event.set()  # notificationflushthreadexit
    flush_thread.join()  # Wait for flush threadcomplete

    # Write remainingoutputwritefile
    captured_output = output_capture.getvalue()
    cleaned_output = clean_output(captured_output)
    if cleaned_output:
        with open(output_file_name, 'a', encoding='utf-8') as f:
            f.write(cleaned_output)

    # close StringIO object
    output_capture.close()

    # Printprocessprogress
    get_rate(output_jsonl)
    print(f"{line_id} completed")


def main():
    # If there is an output file, delete it first
    if os.path.exists(output_name):
        os.remove(output_name)
    if os.path.exists(output_jsonl):
        os.remove(output_jsonl)

    # load reference template and save to new file
    # wb = load_workbook(template_file)
    # wb.save(output_name)

    # Open JSONL file
    with open(input_name, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        if args.n_samples is not None:
            lines = lines[:args.n_samples]
            

    # Use Manager and Pool create process pool
    with Manager() as manager:
        lock = manager.Lock()  # File lock
        pool = Pool(processes=max_workers)

        # Create parameter list to pass to process
        pool_args = [(line, lock) for line in lines]

        # In parallelprocess
        pool.map(process_wrapper, pool_args)

        pool.close()
        pool.join()

if __name__ == "__main__":
    main()
